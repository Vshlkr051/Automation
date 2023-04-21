# Python code to convert excel file from one language to other
# User need to maintain master file where column_1 contains original word and column_2 contains translated word
# ===================================================================================
# Before executing file, user need to install "openpyxl" package into their system
# >>> pip install openpyxl
#  for some user they might have pip3 installed on the system , so they need to run as
# >>> pip3 install openpyxl
# =====================================================================================

import openpyxl

dictionary = input("enter the dictionary file path : ")
file = input("enter the file path to translate : ")
target_file = input("enter the destination path along with file name : ")

# Load "xlsx file 1"
wb1 = openpyxl.load_workbook(dictionary)
sheet1 = wb1.active

# Load "xlsx file 2"
wb2 = openpyxl.load_workbook(file)
sheet2 = wb2.active

# Create a dictionary to store the translation mappings from "xlsx file 1" as key value pair {key : value}
translation_dict = {}
for row in sheet1.iter_rows(values_only=True):
    translation_dict[row[0]] = row[1]

# Translate data in "xlsx file 2" based on the mappings, it will do row by row translation
for row in sheet2.iter_rows():
    for cell in row:
        if cell.value in translation_dict:
            cell.value = translation_dict[cell.value]

# Save the translated data to a new Excel file
wb2.save(target_file)